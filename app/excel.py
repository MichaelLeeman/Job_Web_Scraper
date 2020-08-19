# This module gives functions for setting up the excel worksheet including stylising the table, appending
# the scraped jobs and sorting the job data. It also provides functions for saving and loading workbooks.

from datetime import datetime
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook
from openpyxl import load_workbook


# Updates the worksheet by appending the data, sort it's ordering and calling the other styling functions
def update_xlsx(worksheet, job_list):
    append_jobs_to_xl(job_list, worksheet)
    sort_job_list(worksheet)
    autofit_columns(worksheet)
    colour_rows(worksheet, colour="BBDEFB")
    filter_and_freeze_panes(worksheet)


# Stylise the titles
def create_table_headers(worksheet, font_colour, cell_colour):
    table_headers = ("Job Openings", "Company", "Job Location", "Job Type", "Date Posted", "Deadline", "Salary Range")
    worksheet.append(table_headers)

    for column_title in worksheet[1:1]:
        column_title.font = Font(bold=True, color=font_colour)
        column_title.fill = PatternFill(start_color=cell_colour, fill_type="solid")


# Autofits the columns by taking the length of the longest entry
def autofit_columns(worksheet):
    for column_cell in worksheet.columns:
        max_char_len = 0
        for cell in column_cell:
            if max_char_len < len(str(cell.value)):  # Datetime types need to become strings to measure len
                max_char_len = len(str(cell.value)) * 1.35
        new_column_length = max_char_len
        worksheet.column_dimensions[column_cell[0].column_letter].width = new_column_length


# Colours every other row with the colour parameter
def colour_rows(worksheet, colour):
    for every_other_row in range(2, worksheet.max_row + 1):
        for cell in worksheet[every_other_row]:
            cell.fill = PatternFill(start_color="FFFFFF", fill_type="solid")

    for every_other_row in range(3, worksheet.max_row + 1, 2):
        for cell in worksheet[every_other_row]:
            cell.fill = PatternFill(start_color=colour, fill_type="solid")


# Add filter and freeze pane
def filter_and_freeze_panes(worksheet):
    worksheet.auto_filter.ref = worksheet.dimensions
    freeze_above = worksheet['A2']
    worksheet.freeze_panes = freeze_above
    return worksheet


# Appends each job opening to the worksheet and creates a hyperlink to its page
def append_jobs_to_xl(job_list, worksheet):
    first_xl_job = tuple(cell.value for row in worksheet["A2":"B2"] for cell in row)

    # Fixes bug where a blank row is appended
    if worksheet["A2"].value is None:
        worksheet.delete_rows(2)

    # Only append jobs that are not already in the worksheet.
    for job in job_list:
        # Rather than compare each job in the worksheet, only compare it to the first job to save time
        if job[:2] != first_xl_job:
            worksheet.append(job[0:7])
            current_row = worksheet._current_row

            # Adds a hyperlink to each job web page in the job title column
            worksheet["A" + str(current_row)].hyperlink = job[6]

            # If the company link was given then add the hyperlink
            if job[7] is not None:
                worksheet["B" + str(current_row)].hyperlink = job[7]

        # Stop adding jobs from job_list to the worksheet
        else:
            break


# Removes any job postings that pasted their deadline
def remove_old_jobs(worksheet):
    # Create a list of all the jobs and remove the outdated jobs by comparing the deadline to the current date.
    all_jobs = get_jobs_in_table(worksheet)
    for job in all_jobs:
        if datetime.strptime(job[4], "%d-%b-%Y") < datetime.today():
            all_jobs.remove(job)
    
    # Delete all the jobs currently in the table and append the corrected job list.
    worksheet.delete_rows(2, worksheet.max_row)
    append_jobs_to_xl(all_jobs, worksheet)


# Returns all of the jobs currently in the excel worksheet
def get_jobs_in_table(worksheet):
    table, all_jobs, row = worksheet["A2":"G" + str(worksheet.max_row)], [], 2
    # All the jobs in the table needs to be taken out
    for job in table:
        current_job = []
        for job_detail in job:
            current_job.append(job_detail.value)
        # Try finding the job's and company's link, if the job doesn't have a link then append None onto the end
        try:
            current_job.append(worksheet["A" + str(row)].hyperlink.target)
        except AttributeError:
            current_job.append(None)
        try:
            current_job.append(worksheet["B" + str(row)].hyperlink.target)
        except AttributeError:
            current_job.append(None)
        all_jobs.append(current_job)
        row += 1
    return all_jobs


# Sorts all the jobs in the worksheet, from the most recently posted to the oldest
def sort_job_list(worksheet):
    all_jobs = get_jobs_in_table(worksheet)
    # Delete the current order of the jobs in the worksheet and add the ordered job list
    worksheet.delete_rows(2, worksheet.max_row)
    all_jobs = sorted(all_jobs, key=lambda date: datetime.strptime(date[4], "%d-%b-%Y"), reverse=True)
    append_jobs_to_xl(all_jobs, worksheet)


# Returns the posted date of the first job
def get_first_job_date(worksheet):
    first_job_date = worksheet["E2"].value
    return first_job_date


# Initialise a new workbook and worksheet
def init_xlsx(worksheet_title):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title
    create_table_headers(worksheet, font_colour="FFFFFF", cell_colour="2196F3")
    return workbook, worksheet


# Loads an existing workbook
def load_xlsx(file_path):
    workbook = load_workbook(filename=file_path)
    worksheet = workbook.active

    # Remove the outdated jobs currently in the worksheet
    remove_old_jobs(worksheet)
    return workbook, worksheet


# Saves the workbook
def save_xlsx(workbook, file_path):
    workbook.save(file_path)
