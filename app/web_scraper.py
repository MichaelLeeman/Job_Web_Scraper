# This program scraps data from job postings on the website workinstartups.com and appends it to an excel worksheet.
# So far, the program scraps job titles and company names using BeautifulSoup and adds the data to excel using
# openpyxl.

import time as t
from selenium import webdriver
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup
from datetime import datetime, timedelta

# -----------------------------------------------------------------------
# Web Scraping
# -----------------------------------------------------------------------


# Makes the GET request to the URL links and creates a soup
def soup_creator(URL_link, max_retry=3, sleep_time=0.5):
    # Retry connection if a connection error occurs.
    current_page, request_worked, number_of_total_retries = None, False, 0
    while number_of_total_retries < max_retry and not request_worked:
        try:
            current_page = requests.get(URL_link, headers={"User-Agent": "Chrome/83.0"}, allow_redirects=False)
            request_worked = True
        except requests.exceptions.ConnectionError as err:
            print("Connection error to " + str(URL_link) + " has failed.")
            print("Retrying the connection to the URL attempt number: " + str(number_of_total_retries+1))
            t.sleep((2 ** number_of_total_retries)-1)   # Sleep times [ 0.0s, 1.0s, 3.0s]
            number_of_total_retries += 1
            if number_of_total_retries >= max_retry:
                raise err
    # Creating soup and waiting for elements to load
    current_soup = BeautifulSoup(current_page.text, "html.parser")
    t.sleep(sleep_time)
    return current_soup


# Extracts the job details and hyperlink from each job posting on the current page
def scrape_job_details(soup):
    for div in soup.find_all(name="div", attrs={"class": "job-listing mb-2"}):
        job_hyperlink = div.a["href"]
        job_title = div.a["title"]
        company_name = div.find(name="span", attrs={"style": "display: ruby-base-container"}).string.split(None, 2)[1]
        job_type = div.find("span").string

        # Format the date posted to match other dates
        unformatted_date = div.find("span", attrs={"style": "order: 2"}).string.strip()
        date_posted = datetime.strptime(unformatted_date, "%d-%m-%Y").strftime('%d-%b-%Y')

        # Scrapes for extra details found inside the job's description page (deadline, salary range, etc.).
        job_description_soup = soup_creator(job_hyperlink)

        # Deadline is written in a string text. To extract the date, the text is converted to a list.
        date_text_into_list = job_description_soup.find("small").string.split()[8: 11]
        separator = '-'
        expiry_date = separator.join(date_text_into_list)  # Joining only the day, month and year back into a string

        # Sometimes the salary range might be given in the salary icon, others specified in the text. If not then assign "Unspecified salary".
        salary_contents = job_description_soup.find(attrs={"class": "mb-3 mb-sm-0"})

        if salary_contents is not None:
            # If salary is given in the icon then the salary range can be scraped from it like a list
            salary_range = salary_contents.contents[1].strip()
        else:
            salary_range, commission_already_added, equity_already_added = "Unspecified salary", False, False

            # Unpaid positions, commission only and other salary types are only specified in the job description text
            for p_element in job_description_soup.find_all(name="p"):
                job_description_text = p_element.text.lower()

                if "unpaid" in job_description_text or "voluntary" in job_description_text or "volunteer" in job_description_text or "no salary" in job_description_text:
                    salary_range = "Unpaid"

                elif "competitive salary" in job_description_text:
                    salary_range = "Competitive salary"

                # Some jobs have commission with other salary types. Others have only commission.
                if "commission" in job_description_text:
                    if not commission_already_added:
                        salary_range += " + commission"
                        commission_already_added = True
                    commission_texts = ["commission only", "commission-only", "only commission", "commission based"]
                    for commission in commission_texts:
                        if commission in job_description_text:
                            salary_range = "Commission only"
                            break

                if "equity" in job_description_text and "private equity" not in job_description_text and not equity_already_added:
                    salary_range += " + equity"
                    equity_already_added = True

        # Scraping the hyperlink to the company's website
        company_hyperlink_element = job_description_soup.find(attrs={"class": "d-flex my-4 container"})
        company_hyperlink = company_hyperlink_element.a["href"]

        # Sometimes company URLs aren't given which means the next found element with attribute href is a link to another page
        if "https://workinstartups.com/" in company_hyperlink:
            company_hyperlink = None

        hyperlink_list.append(job_hyperlink)
        company_link_list.append(company_hyperlink)
        job_list.append((job_title, company_name, job_type, date_posted, expiry_date, salary_range))

    return job_list, hyperlink_list, company_link_list


# Checks the posted date of the last job to see whether to keep searching for recently posted jobs.
def check_date(job_list, last_date):
    last_job_datetime = datetime.strptime(job_list[-1][3], '%d-%b-%Y')  # Needs to convert back to datetime to make comparison
    if last_job_datetime < last_date:
        keep_searching = False
    else:
        keep_searching = True
    return keep_searching


# Removes jobs from the list if they where posted kore than a fortnight ago
def remove_outdated_jobs(job_list, last_date):
    for job in job_list[:]:
        job_datetime = datetime.strptime(job[3], '%d-%b-%Y')
        if job_datetime < last_date:
            job_list.remove(job)
    return job_list


# Goes to the next page
def go_to_new_page():
    driver.find_element_by_link_text('Next >').click()
    new_soup = soup_creator(driver.current_url)
    return new_soup


# Keeps scraping for jobs on the current page while checking that they aren't older than a fortnight ago.
def search_for_jobs(current_soup, last_date_to_check):
    keep_searching_for_jobs = True
    while keep_searching_for_jobs:
        unsorted_job_list, job_hyperlink_list, company_hyperlink_list = scrape_job_details(current_soup)
        keep_searching_for_jobs = check_date(job_list, last_date_to_check)
        current_soup = go_to_new_page()
    sorted_job_list = remove_outdated_jobs(job_list, last_date_to_check)
    return sorted_job_list, job_hyperlink_list, company_hyperlink_list


URL = "https://workinstartups.com/job-board/jobs-in/london"
soup = soup_creator(URL, max_retry=1, sleep_time=0)

driver = webdriver.Chrome('./chromedriver')
driver.get(URL)
driver.find_element_by_link_text('Close').click()

current_date = datetime.today()
date_fortnight_ago = current_date - timedelta(weeks=2)
last_recent_date = date_fortnight_ago.replace(hour=0, minute=0, second=0, microsecond=0)  # default to midnight

job_list, hyperlink_list, company_link_list = [], [], []
job_list, hyperlink_list, company_link_list = search_for_jobs(soup, last_recent_date)
driver.close()

# -----------------------------------------------------------------------
# Excel
# -----------------------------------------------------------------------


def setup_worksheet(worksheet):
    title_names = ("Job Openings", "Company", "Job Type", "Date Posted", "Deadline", "Salary Range")
    worksheet.append(title_names)

    # Stylise the titles
    for column_title in worksheet[1:1]:
        column_title.font = Font(bold=True, color='FFFFFF')
        column_title.fill = PatternFill(start_color="2196F3", fill_type="solid")

    append_job_to_xl(job_list, worksheet)

    # Autofits the columns by taking the length of the longest entry
    for column_cell in worksheet.columns:
        max_char_len = 0
        for cell in column_cell:
            if max_char_len < len(str(cell.value)):  # Datetime types need to become strings to measure len
                max_char_len = len(str(cell.value))
        new_column_length = max_char_len
        worksheet.column_dimensions[column_cell[0].column_letter].width = new_column_length

    # Colours every other row light blue
    for every_other_row in range(3, worksheet.max_row + 1, 2):
        for cell in worksheet[every_other_row]:
            cell.fill = PatternFill(start_color="BBDEFB", fill_type="solid")

    # Add filter and freeze pane
    worksheet.auto_filter.ref = worksheet.dimensions
    freeze_above = worksheet['A2']
    worksheet.freeze_panes = freeze_above


# Appends each job opening to the worksheet and creates a hyperlink to its page
def append_job_to_xl(job_list, worksheet):
    URL_index = 0
    for job in job_list:
        worksheet.append(job)
        current_row = worksheet._current_row
        # Adds a hyperlink to each job web page in the job title column
        worksheet["A" + str(current_row)].hyperlink = hyperlink_list[URL_index]
        if company_link_list[URL_index] is not None:
            worksheet["B" + str(current_row)].hyperlink = company_link_list[URL_index]
        URL_index += 1


file_path = "Job_Openings.xlsx"
book = Workbook()
sheet1 = book.active
setup_worksheet(sheet1)
book.save(file_path)
