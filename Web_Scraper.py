# This program scraps data from job postings on the website workinstartups.com and appends it to an excel worksheet.
# So far, the program scraps job titles and company names using BeautifulSoup and adds the data to excel using
# openpyxl.

import time as t
from selenium import webdriver
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from bs4 import BeautifulSoup
from datetime import *

# -----------------------------------------------------------------------
# Web Scraping
# -----------------------------------------------------------------------

URL = "https://workinstartups.com/job-board/jobs-in/london"
page = requests.get(URL)
soup = BeautifulSoup(page.text, "html.parser")

driver = webdriver.Chrome('./chromedriver')
driver.get("https://workinstartups.com/job-board/jobs-in/london")
driver.find_element_by_link_text('Close').click()

job_list = []
last_recent_date = datetime.strptime("19-05-2020", "%d-%m-%Y")


# Extracts the job details from each job posting on the current page.
def scrape_job_details(soup):
    for div in soup.find_all(name="div", attrs={"class": "job-listing mb-2"}):
        job_title = div.a["title"]
        company_name = div.find(name="span", attrs={"style": "display: ruby-base-container"}).string.split(None, 2)[1]
        job_type = div.find("span").string
        date_posted = div.find("span", attrs={"style": "order: 2"}).string.strip()
        job_list.append((job_title, company_name, job_type, date_posted))
    return job_list


# Checks whether the date posted of every job and removes it from the list if it's too old.
# Returns a boolean to stop searching for jobs once they are old ones.
def remove_outdated_jobs(job_list, keep_searching):
    for job in job_list[:]:
        job_datetime = datetime.strptime(job[3], "%d-%m-%Y")
        if job_datetime < last_recent_date:
            job_list.remove(job)
            keep_searching = False
    return job_list, keep_searching


# Goes to the next page
def go_to_new_page():
    driver.find_element_by_link_text('Next >').click()
    current_page = requests.get(driver.current_url)
    new_soup = BeautifulSoup(current_page.text, "html.parser")
    return new_soup


# Keep adding the jobs from the page until they are older than than "last_recent_date"
def search_for_jobs(current_soup, job_list):
    keep_searching_for_jobs = True
    while keep_searching_for_jobs:
        job_list, keep_searching_for_jobs = remove_outdated_jobs(scrape_job_details(current_soup), keep_searching_for_jobs)
        t.sleep(1)
        current_soup = go_to_new_page()
        t.sleep(1)
    return job_list


job_list = search_for_jobs(soup, job_list)
driver.close()

# -----------------------------------------------------------------------
# Excel
# -----------------------------------------------------------------------


def setup_worksheet(worksheet):
    title_names = ("Job Openings", "Company", "Job Type", "Date Posted")
    worksheet.append(title_names)

    # Stylise the titles
    for column_title in worksheet[1:1]:
        column_title.font = Font(bold=True, color='FFFFFF')
        column_title.fill = PatternFill(start_color="2196F3", fill_type="solid")

    append_job_to_xl(job_list, worksheet)

    # Autofits the columns by taking the length of the longest entry
    for column_cell in worksheet.columns:
        max_char_len = 0
        for cell in column_cell:
            if max_char_len < len(cell.value):
                max_char_len = len(cell.value)
        new_column_length = max_char_len * 0.95
        worksheet.column_dimensions[column_cell[0].column_letter].width = new_column_length

    # Colours every other row blue
    for every_other_row in range(3, worksheet.max_row + 1, 2):
        for cell in worksheet[every_other_row]:
            cell.fill = PatternFill(start_color="BBDEFB", fill_type="solid")


# Appends each job opening to the worksheet.
def append_job_to_xl(job_list, worksheet):
    for job in job_list:
        worksheet.append(job)


file_path = "Job_Openings.xlsx"
book = Workbook()
sheet1 = book.active
setup_worksheet(sheet1)
book.save(file_path)